from flask import Flask, render_template, request, redirect, flash, session, url_for, jsonify
from database import get_connection, log_activity, create_notification
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from datetime import timedelta
import os
import time
import re

app = Flask(__name__)
app.secret_key = "student_management_secret_erp"
app.permanent_session_lifetime = timedelta(minutes=30)

# Config uploads
UPLOAD_FOLDER = os.path.join('static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2MB limits
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==========================================
# RBAC DECORATORS
# ==========================================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash("Access denied. Please log in first.", "warning")
            return redirect(url_for('login'))
        
        # Verify account status is still active in the database
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT status, role FROM users WHERE id = %s", (session['user']['id'],))
        res = cur.fetchone()
        cur.close()
        conn.close()
        
        if not res or res[0] != 'Active':
            session.pop('user', None)
            flash("Your account has been deactivated or suspended.", "danger")
            return redirect(url_for('login'))
            
        # Update session role just in case admin changed it
        session['user']['role'] = res[1]
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session['user'].get('role') != 'Admin':
            flash("Access denied. Admin privileges required.", "danger")
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

def hod_or_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session['user'].get('role') not in ['Admin', 'HOD']:
            flash("Access denied. HOD or Admin privileges required.", "danger")
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

# Context processor to inject notifications into all templates
@app.context_processor
def inject_notifications():
    if 'user' in session:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, type, message, is_read, created_at 
            FROM notifications 
            WHERE user_id = %s 
            ORDER BY created_at DESC LIMIT 15
        """, (session['user']['id'],))
        notifs = cur.fetchall()
        
        cur.execute("SELECT COUNT(*) FROM notifications WHERE user_id = %s AND is_read = FALSE", (session['user']['id'],))
        unread_count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return dict(notifications=notifs, unread_notifications_count=unread_count)
    return dict(notifications=[], unread_notifications_count=0)

# ==========================================
# AUTHENTICATION ROUTES
# ==========================================
@app.route("/login", methods=["GET", "POST"])
def login():
    if 'user' in session:
        return redirect(url_for('home'))
        
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cur.fetchall()
    cur.close()
    conn.close()

    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, username, password_hash, status, role, full_name 
            FROM users WHERE username = %s
        """, (username,))
        user_record = cur.fetchone()
        
        if user_record:
            uid, uname, pw_hash, status, role, full_name = user_record
            if check_password_hash(pw_hash, password):
                if status != 'Active':
                    cur.close()
                    conn.close()
                    if status == 'Pending':
                        flash("Your request has been submitted successfully and is awaiting administrator approval.", "info")
                    elif status == 'Suspended':
                        flash("Your account has been suspended. Please contact the administrator.", "danger")
                    else:
                        flash("Your account request was rejected.", "danger")
                    return redirect(url_for('login'))
                
                # Active user login
                session.permanent = True
                session['user'] = {
                    'id': uid,
                    'username': uname,
                    'role': role,
                    'full_name': full_name
                }
                
                # Update last login timestamp
                cur.execute("UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = %s", (uid,))
                conn.commit()
                cur.close()
                conn.close()
                
                log_activity("Faculty Login", f"User '{username}' logged in successfully.", username)
                
                # Notification alert
                create_notification(uid, 'Security', f"Logged in from IP: {request.remote_addr}")
                
                flash("Welcome back! Logged in successfully.", "success")
                return redirect(url_for('home'))
        
        cur.close()
        conn.close()
        log_activity("Failed Login Attempt", f"Username attempted: '{username}'")
        flash("Invalid username or password.", "danger")
            
    return render_template("login.html", departments=departments)

@app.route("/register", methods=["POST"])
def register():
    full_name = request.form["full_name"].strip()
    employee_id = request.form["employee_id"].strip().upper()
    email = request.form["email"].strip().lower()
    department_id = int(request.form["department_id"])
    designation = request.form["designation"].strip()
    username = request.form["username"].strip()
    password = request.form["password"].strip()
    confirm_password = request.form["confirm_password"].strip()

    if password != confirm_password:
        flash("Passwords do not match.", "danger")
        return redirect(url_for('login'))

    # Validate secure username/password rules
    is_valid_uname, error_uname = is_secure_username(username)
    if not is_valid_uname:
        flash(error_uname, "danger")
        return redirect(url_for('login'))

    is_valid_pwd, error_pwd = is_secure_password(password)
    if not is_valid_pwd:
        flash(error_pwd, "danger")
        return redirect(url_for('login'))

    # Handle optional Faculty ID file upload
    image_path = None
    if 'faculty_id_card' in request.files:
        file = request.files['faculty_id_card']
        if file and file.filename != '' and allowed_file(file.filename):
            filename = f"fac_id_{employee_id}_{int(time.time())}_{secure_filename(file.filename)}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_path = f"static/uploads/{filename}"

    conn = get_connection()
    cur = conn.cursor()

    try:
        # Check duplicates
        cur.execute("SELECT COUNT(*) FROM users WHERE username = %s", (username,))
        if cur.fetchone()[0] > 0:
            flash("Username is already taken.", "warning")
            return redirect(url_for('login'))

        cur.execute("SELECT COUNT(*) FROM users WHERE email = %s", (email,))
        if cur.fetchone()[0] > 0:
            flash("Email address is already registered.", "warning")
            return redirect(url_for('login'))

        cur.execute("SELECT COUNT(*) FROM users WHERE employee_id = %s", (employee_id,))
        if cur.fetchone()[0] > 0:
            flash("Employee ID is already registered.", "warning")
            return redirect(url_for('login'))

        hashed_pwd = generate_password_hash(password)
        cur.execute("""
            INSERT INTO users (username, password_hash, email, full_name, employee_id, department_id, designation, image_path, status, role)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Pending', 'Faculty')
            RETURNING id;
        """, (username, hashed_pwd, email, full_name, employee_id, department_id, designation, image_path))
        new_uid = cur.fetchone()[0]
        conn.commit()

        # Notify Administrators
        cur.execute("SELECT id FROM users WHERE role = 'Admin'")
        admin_ids = [row[0] for row in cur.fetchall()]
        for aid in admin_ids:
            create_notification(aid, 'System Alert', f"New account request submitted by {full_name} ({employee_id})")

        log_activity("Faculty Account Requested", f"User '{username}' registered with status Pending.", username)
        flash("Your request has been submitted successfully and is awaiting administrator approval.", "info")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to register account: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('login'))

@app.route("/forgot-password", methods=["POST"])
def forgot_password():
    employee_id = request.form["employee_id"].strip().upper()
    email = request.form["email"].strip().lower()
    new_password = request.form["new_password"].strip()
    confirm_password = request.form["confirm_password"].strip()

    if new_password != confirm_password:
        flash("Passwords do not match.", "danger")
        return redirect(url_for('login'))

    is_valid_pwd, error_pwd = is_secure_password(new_password)
    if not is_valid_pwd:
        flash(error_pwd, "danger")
        return redirect(url_for('login'))

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, username FROM users WHERE employee_id = %s AND email = %s", (employee_id, email))
    user_record = cur.fetchone()

    if not user_record:
        cur.close()
        conn.close()
        flash("Security verification failed. Employee ID and Email do not match.", "danger")
        return redirect(url_for('login'))

    uid, username = user_record
    hashed_pwd = generate_password_hash(new_password)
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (hashed_pwd, uid))
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Password Recovered", f"User '{username}' recovered password using Employee ID verification.", username)
    flash("Password reset successfully. Please log in with your new password.", "success")
    return redirect(url_for('login'))

@app.route("/logout")
def logout():
    if 'user' in session:
        username = session['user']['username']
        session.pop('user', None)
        log_activity("Faculty Logout", f"User '{username}' logged out.", username)
        flash("You have been logged out.", "info")
    return redirect(url_for('login'))

# ==========================================
# HOME / READ (DASHBOARD & REGISTRY)
# ==========================================
@app.route("/")
@login_required
def home():
    username = session['user']['username']
    user_role = session['user']['role']
    search = request.args.get("search", "")
    department_filter = request.args.get("department", "")
    status_filter = request.args.get("status", "")
    gpa_min = request.args.get("gpa_min", "")
    gpa_max = request.args.get("gpa_max", "")
    age_filter = request.args.get("age", "")
    timeline_filter = request.args.get("timeline_filter", "Today")

    conn = get_connection()
    cur = conn.cursor()

    # ---------------- Dashboard Stats ----------------
    cur.execute("SELECT COUNT(*) FROM students")
    total_students = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM departments")
    total_departments = cur.fetchone()[0]

    cur.execute("SELECT ROUND(AVG(gpa), 2) FROM students")
    avg_gpa_val = cur.fetchone()[0]
    avg_gpa = float(avg_gpa_val) if avg_gpa_val is not None else 0.00

    cur.execute("SELECT COUNT(*) FROM students WHERE status='Active'")
    active_students = cur.fetchone()[0]

    # Additional stats
    cur.execute("SELECT name, gpa FROM students ORDER BY gpa DESC, name LIMIT 1")
    top_perf = cur.fetchone()
    top_performer = f"{top_perf[0]} ({top_perf[1]})" if top_perf else "N/A"

    cur.execute("SELECT name, gpa FROM students ORDER BY gpa ASC, name LIMIT 1")
    low_perf = cur.fetchone()
    lowest_gpa = f"{low_perf[0]} ({low_perf[1]})" if low_perf else "N/A"

    cur.execute("SELECT name FROM students ORDER BY created_at DESC LIMIT 1")
    recent_student = cur.fetchone()
    recently_added = recent_student[0] if recent_student else "N/A"

    cur.execute("SELECT COUNT(*) FROM students WHERE status = 'Inactive'")
    inactive_students = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM students WHERE status = 'Graduated'")
    graduated_students = cur.fetchone()[0]

    # ---------------- Faculty KPIs (Admin/HOD view) ----------------
    faculty_kpi = {}
    if user_role in ['Admin', 'HOD']:
        cur.execute("SELECT COUNT(*) FROM users")
        faculty_kpi['total'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM users WHERE status = 'Pending'")
        faculty_kpi['pending'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM users WHERE status = 'Active'")
        faculty_kpi['active'] = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM users WHERE status = 'Suspended'")
        faculty_kpi['suspended'] = cur.fetchone()[0]

    # ---------------- AI Academic Insights ----------------
    insights = []
    if total_students > 0:
        # Highest performing dept
        cur.execute("""
            SELECT d.name, ROUND(COALESCE(AVG(s.gpa), 0), 2) AS avg_gpa, COUNT(s.id)
            FROM departments d
            JOIN students s ON s.department_id = d.id
            GROUP BY d.id, d.name
            ORDER BY avg_gpa DESC LIMIT 1
        """)
        high_dept = cur.fetchone()
        if high_dept:
            insights.append({
                'title': 'Highest Performing Department',
                'desc': f"{high_dept[0]} leads academically with an average GPA of {high_dept[1]} across {high_dept[2]} students.",
                'icon': 'trending_up',
                'class': 'success'
            })
            
        # Lowest performing dept
        cur.execute("""
            SELECT d.name, ROUND(COALESCE(AVG(s.gpa), 0), 2) AS avg_gpa
            FROM departments d
            JOIN students s ON s.department_id = d.id
            GROUP BY d.id, d.name
            ORDER BY avg_gpa ASC LIMIT 1
        """)
        low_dept = cur.fetchone()
        if low_dept and (not high_dept or low_dept[0] != high_dept[0]):
            insights.append({
                'title': 'Targeted Support Needed',
                'desc': f"{low_dept[0]} has the lowest average GPA of {low_dept[1]}. Recommend academic mentoring interventions.",
                'icon': 'announcement',
                'class': 'warning'
            })

    # ---------------- Risk Prediction Calculations ----------------
    cur.execute("""
        SELECT s.id, s.name, s.gpa, s.status, d.name, s.age 
        FROM students s
        LEFT JOIN departments d ON s.department_id = d.id
        ORDER BY s.gpa ASC
    """)
    all_students_for_risk = cur.fetchall()
    at_risk_list = []
    high_risk_count = 0
    medium_risk_count = 0
    
    for row in all_students_for_risk:
        gpa_val = float(row[2])
        status_val = row[3]
        
        if status_val == 'Inactive':
            level = 'High'
            reason = 'Enrollment status is Inactive'
            high_risk_count += 1
        elif gpa_val < 6.0:
            level = 'High'
            reason = 'Critical GPA (< 6.00)'
            high_risk_count += 1
        elif gpa_val < 7.5:
            level = 'Medium'
            reason = 'Below average GPA (< 7.50)'
            medium_risk_count += 1
        else:
            level = 'Low'
            reason = 'Academic performance stable'
            
        if level in ['High', 'Medium']:
            at_risk_list.append({
                'id': row[0],
                'name': row[1],
                'gpa': gpa_val,
                'status': status_val,
                'department': row[4] or 'Unassigned',
                'level': level,
                'reason': reason
            })
    
    # Sort: High risk first
    at_risk_list.sort(key=lambda x: (0 if x['level'] == 'High' else 1, x['gpa']))

    # ---------------- Department Leaderboard ----------------
    cur.execute("""
        SELECT d.id, d.name, d.code, 
               COUNT(s.id) AS student_count, 
               ROUND(COALESCE(AVG(s.gpa), 0), 2) AS avg_gpa,
               COALESCE(MAX(s.gpa), 0) AS max_gpa
        FROM departments d
        LEFT JOIN students s ON s.department_id = d.id
        GROUP BY d.id, d.name, d.code
        ORDER BY avg_gpa DESC
    """)
    leaderboard = []
    for rank, row in enumerate(cur.fetchall(), 1):
        leaderboard.append({
            'rank': rank,
            'id': row[0],
            'name': row[1],
            'code': row[2],
            'student_count': row[3],
            'avg_gpa': float(row[4]),
            'max_gpa': float(row[5]),
            'score': round((float(row[4]) / 10.00) * 100, 1)
        })

    # ---------------- Analytics Charts ----------------
    cur.execute("""
        SELECT d.name, COUNT(s.id), ROUND(COALESCE(AVG(s.gpa), 0), 2)
        FROM departments d
        LEFT JOIN students s ON s.department_id = d.id
        GROUP BY d.id, d.name
        ORDER BY d.name
    """)
    dept_stats = cur.fetchall()

    cur.execute("""
        SELECT 
            COUNT(CASE WHEN gpa < 5 THEN 1 END) as gpa_0_5,
            COUNT(CASE WHEN gpa >= 5 AND gpa < 7.5 THEN 1 END) as gpa_5_7_5,
            COUNT(CASE WHEN gpa >= 7.5 AND gpa < 8.5 THEN 1 END) as gpa_7_5_8_5,
            COUNT(CASE WHEN gpa >= 8.5 THEN 1 END) as gpa_8_5_10
       FROM students
    """)
    gpa_dist = list(cur.fetchone())

    cur.execute("SELECT status, COUNT(*) FROM students GROUP BY status")
    status_counts = dict(cur.fetchall())
    status_dist = {
        'Active': status_counts.get('Active', 0),
        'Inactive': status_counts.get('Inactive', 0),
        'Graduated': status_counts.get('Graduated', 0)
    }

    cur.execute("""
        SELECT 
            COUNT(CASE WHEN age < 18 THEN 1 END) as age_under_18,
            COUNT(CASE WHEN age >= 18 AND age <= 21 THEN 1 END) as age_18_21,
            COUNT(CASE WHEN age >= 22 AND age <= 25 THEN 1 END) as age_22_25,
            COUNT(CASE WHEN age > 25 THEN 1 END) as age_above_25
        FROM students
    """)
    age_dist = list(cur.fetchone())

    cur.execute("""
        SELECT TO_CHAR(created_at, 'YYYY-MM') AS month, COUNT(*) 
        FROM students 
        GROUP BY month 
        ORDER BY month
    """)
    monthly_data = cur.fetchall()
    months = [row[0] for row in monthly_data]
    monthly_counts = [row[1] for row in monthly_data]
    
    cumulative_counts = []
    total = 0
    for count in monthly_counts:
        total += count
        cumulative_counts.append(total)

    analytics_data = {
        'departments': [row[0] for row in dept_stats],
        'studentCounts': [row[1] for row in dept_stats],
        'avgGpas': [float(row[2]) for row in dept_stats],
        'gpaDist': gpa_dist,
        'statusDist': status_dist,
        'ageDist': age_dist,
        'months': months,
        'monthlyCounts': monthly_counts,
        'cumulativeCounts': cumulative_counts
    }

    # Departments for Filter / Form
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cur.fetchall()

    # Recent Announcements
    cur.execute("""
        SELECT a.id, a.title, a.description, a.category, a.priority, a.created_at, u.username
        FROM announcements a
        LEFT JOIN users u ON a.author_id = u.id
        WHERE a.scheduled_date <= CURRENT_TIMESTAMP 
          AND (a.expiry_date IS NULL OR a.expiry_date >= CURRENT_TIMESTAMP)
        ORDER BY a.created_at DESC LIMIT 5
    """)
    announcements_list = cur.fetchall()

    # Students Records (Filtered)
    query = """
        SELECT s.id, s.name, s.age, d.name AS department_name, s.email, s.gpa, s.status, s.image_path, s.department_id
        FROM students s
        LEFT JOIN departments d ON s.department_id = d.id
        WHERE s.name ILIKE %s
    """
    values = [f"%{search}%"]
    
    if department_filter:
        query += " AND s.department_id = %s"
        values.append(int(department_filter))
    if status_filter:
        query += " AND s.status = %s"
        values.append(status_filter)
    if gpa_min:
        query += " AND s.gpa >= %s"
        values.append(float(gpa_min))
    if gpa_max:
        query += " AND s.gpa <= %s"
        values.append(float(gpa_max))
    if age_filter:
        query += " AND s.age = %s"
        values.append(int(age_filter))

    query += " ORDER BY s.id"
    cur.execute(query, tuple(values))
    students = cur.fetchall()

    # Faculty Activity Logs
    timeline_query = """
        SELECT action, details, created_at, faculty_username 
        FROM activity_logs 
    """
    timeline_vals = []
    if timeline_filter == "Today":
        timeline_query += " WHERE created_at >= CURRENT_DATE"
    elif timeline_filter == "This Week":
        timeline_query += " WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'"
    elif timeline_filter == "This Month":
        timeline_query += " WHERE created_at >= CURRENT_DATE - INTERVAL '30 days'"
        
    timeline_query += " ORDER BY created_at DESC LIMIT 30"
    cur.execute(timeline_query, tuple(timeline_vals))
    activities = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "index.html",
        students=students,
        total_students=total_students,
        total_departments=total_departments,
        avg_gpa=avg_gpa,
        active_students=active_students,
        inactive_students=inactive_students,
        graduated_students=graduated_students,
        top_performer=top_performer,
        lowest_gpa=lowest_gpa,
        recently_added=recently_added,
        dept_stats=dept_stats,
        departments=departments,
        activities=activities,
        analytics_data=analytics_data,
        announcements=announcements_list,
        insights=insights,
        at_risk_list=at_risk_list[:5],
        high_risk_count=high_risk_count,
        medium_risk_count=medium_risk_count,
        leaderboard=leaderboard[:3],
        faculty_kpi=faculty_kpi,
        selected_department=department_filter,
        selected_status=status_filter,
        gpa_min=gpa_min,
        gpa_max=gpa_max,
        age_filter=age_filter,
        timeline_filter=timeline_filter,
        search=search
    )

# ==========================================
# STUDENT PROFILE VIEW (WITH NOTES & TIMELINES)
# ==========================================
@app.route("/student/<int:id>")
@login_required
def student_profile(id):
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT s.id, s.name, s.age, d.name AS department_name, s.email, s.gpa, s.status, s.image_path, s.created_at
        FROM students s
        LEFT JOIN departments d ON s.department_id = d.id
        WHERE s.id = %s
    """, (id,))
    student = cur.fetchone()
    
    if not student:
        cur.close()
        conn.close()
        flash("Student profile not found.", "warning")
        return redirect(url_for('home'))

    # Notes
    cur.execute("""
        SELECT n.id, n.note, n.created_at, u.username
        FROM student_notes n
        LEFT JOIN users u ON n.faculty_id = u.id
        WHERE n.student_id = %s
        ORDER BY n.created_at DESC
    """, (id,))
    notes = cur.fetchall()

    # Timeline
    cur.execute("""
        SELECT event_type, description, created_at 
        FROM student_timeline 
        WHERE student_id = %s
        ORDER BY created_at DESC
    """, (id,))
    timeline = cur.fetchall()
    
    cur.close()
    conn.close()
    
    log_activity("Profile Viewed", f"Viewed profile of student {student[1]} (ID: {id})", username)
    return render_template("profile.html", student=student, notes=notes, timeline=timeline)

# ==========================================
# PRIVATE NOTES ENDPOINTS
# ==========================================
@app.route("/student/<int:id>/notes/add", methods=["POST"])
@login_required
def add_student_note(id):
    username = session['user']['username']
    faculty_id = session['user']['id']
    note_text = request.form["note"].strip()

    if not note_text:
        flash("Cannot save an empty note.", "warning")
        return redirect(url_for('student_profile', id=id))

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO student_notes (student_id, faculty_id, note)
        VALUES (%s, %s, %s)
    """, (id, faculty_id, note_text))
    
    cur.execute("""
        INSERT INTO student_timeline (student_id, event_type, description)
        VALUES (%s, 'Faculty Remark', %s)
    """, (id, f"Note added by Faculty '{username}': {note_text[:50]}..."))
    
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Student Note Added", f"Added note to student ID: {id}", username)
    flash("Private note added successfully.", "success")
    return redirect(url_for('student_profile', id=id))

@app.route("/student/<int:sid>/notes/delete/<int:nid>")
@login_required
def delete_student_note(sid, nid):
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM student_notes WHERE id = %s AND student_id = %s", (nid, sid))
    
    cur.execute("""
        INSERT INTO student_timeline (student_id, event_type, description)
        VALUES (%s, 'Profile Modification', %s)
    """, (sid, f"Faculty note deleted by '{username}'"))
    
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Student Note Deleted", f"Deleted note ID {nid} for student ID {sid}", username)
    flash("Private note deleted.", "info")
    return redirect(url_for('student_profile', id=sid))

# ==========================================
# ADD STUDENT (WITH AUTO TIMELINE)
# ==========================================
@app.route("/add", methods=["POST"])
@login_required
def add_student():
    username = session['user']['username']
    name = request.form["name"].strip()
    age = int(request.form["age"])
    department_id = int(request.form["department_id"])
    email = request.form.get("email", "").strip()
    gpa = float(request.form.get("gpa", "0.00"))
    status = request.form.get("status", "Active").strip()

    image_path = None
    if 'profile_image' in request.files:
        file = request.files['profile_image']
        if file and file.filename != '' and allowed_file(file.filename):
            filename = f"avatar_{int(time.time())}_{secure_filename(file.filename)}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_path = f"static/uploads/{filename}"

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM students WHERE email = %s", (email,))
    if cur.fetchone()[0] > 0:
        cur.close()
        conn.close()
        flash(f"Failed to add student. A record with email '{email}' already exists!", "warning")
        return redirect(url_for('home'))

    cur.execute("""
        INSERT INTO students (name, age, department_id, email, gpa, status, image_path)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id;
    """, (name, age, department_id, email, gpa, status, image_path))
    new_id = cur.fetchone()[0]

    cur.execute("""
        INSERT INTO student_timeline (student_id, event_type, description)
        VALUES (%s, 'Admission', %s)
    """, (new_id, f"Admitted to university with status '{status}' and initial GPA of {gpa}"))
    conn.commit()
    
    cur.execute("SELECT id FROM users")
    faculty_ids = [row[0] for row in cur.fetchall()]
    cur.close()
    conn.close()

    for fid in faculty_ids:
        create_notification(fid, 'Student Alert', f"New student registered: {name}")

    log_activity("Student Added", f"Added student: {name} (GPA: {gpa})", username)
    flash(f"Student '{name}' added successfully!", "success")
    return redirect(url_for('home'))

# ==========================================
# UPDATE STUDENT (WITH AUTO TIMELINE)
# ==========================================
@app.route("/update/<int:id>", methods=["POST"])
@login_required
def update_student(id):
    username = session['user']['username']
    name = request.form["name"].strip()
    age = int(request.form["age"])
    department_id = int(request.form["department_id"])
    email = request.form.get("email", "").strip()
    gpa = float(request.form.get("gpa", "0.00"))
    status = request.form.get("status", "Active").strip()

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT name, age, department_id, email, gpa, status FROM students WHERE id = %s", (id,))
    old = cur.fetchone()
    
    if not old:
        cur.close()
        conn.close()
        flash("Student not found.", "warning")
        return redirect(url_for('home'))

    cur.execute("SELECT COUNT(*) FROM students WHERE email = %s AND id != %s", (email, id))
    if cur.fetchone()[0] > 0:
        cur.close()
        conn.close()
        flash(f"Failed to update student. A record with email '{email}' already exists!", "warning")
        return redirect(url_for('home'))

    image_path = None
    if 'profile_image' in request.files:
        file = request.files['profile_image']
        if file and file.filename != '' and allowed_file(file.filename):
            filename = f"avatar_{int(time.time())}_{secure_filename(file.filename)}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            image_path = f"static/uploads/{filename}"

    if image_path:
        cur.execute("""
            UPDATE students
            SET name=%s, age=%s, department_id=%s, email=%s, gpa=%s, status=%s, image_path=%s
            WHERE id=%s
        """, (name, age, department_id, email, gpa, status, image_path, id))
    else:
        cur.execute("""
            UPDATE students
            SET name=%s, age=%s, department_id=%s, email=%s, gpa=%s, status=%s
            WHERE id=%s
        """, (name, age, department_id, email, gpa, status, id))

    # Log changes
    if old[0] != name:
        cur.execute("INSERT INTO student_timeline (student_id, event_type, description) VALUES (%s, 'Profile Modification', %s)", (id, f"Name changed from '{old[0]}' to '{name}'"))
    if old[2] != department_id:
        cur.execute("SELECT name FROM departments WHERE id = %s", (department_id,))
        new_dept_name = cur.fetchone()[0]
        cur.execute("INSERT INTO student_timeline (student_id, event_type, description) VALUES (%s, 'Department Change', %s)", (id, f"Transferred department to '{new_dept_name}'"))
    if float(old[4]) != gpa:
        cur.execute("INSERT INTO student_timeline (student_id, event_type, description) VALUES (%s, 'GPA Update', %s)", (id, f"GPA altered from {old[4]} to {gpa}"))
    if old[5] != status:
        cur.execute("INSERT INTO student_timeline (student_id, event_type, description) VALUES (%s, 'Status Update', %s)", (id, f"Enrollment status changed from '{old[5]}' to '{status}'"))

    conn.commit()
    cur.close()
    conn.close()
    
    log_activity("Student Updated", f"Updated student profile: {name} (ID: {id})", username)
    flash(f"Student '{name}' updated successfully!", "info")
    return redirect(url_for('home'))

# ==========================================
# DELETE STUDENT
# ==========================================
@app.route("/delete/<int:id>")
@login_required
def delete_student(id):
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT name FROM students WHERE id=%s", (id,))
    res = cur.fetchone()
    name = res[0] if res else f"ID: {id}"
    
    cur.execute("DELETE FROM students WHERE id=%s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    
    log_activity("Student Deleted", f"Deleted student: {name} (ID: {id})", username)
    flash("Student deleted successfully!", "danger")
    return redirect(url_for('home'))

# ==========================================
# BULK OPERATIONS ROUTE
# ==========================================
@app.route("/bulk", methods=["POST"])
@login_required
def bulk_operations():
    username = session['user']['username']
    student_ids = request.form.getlist("student_ids")
    action = request.form.get("action")
    status_val = request.form.get("status_value")

    if not student_ids:
        flash("No students selected for bulk operation.", "warning")
        return redirect(url_for('home'))

    ids = [int(sid) for sid in student_ids]
    conn = get_connection()
    cur = conn.cursor()

    if action == "delete":
        cur.execute("SELECT name FROM students WHERE id = ANY(%s)", (ids,))
        names = ", ".join([row[0] for row in cur.fetchall()])
        cur.execute("DELETE FROM students WHERE id = ANY(%s)", (ids,))
        log_activity("Bulk Student Deletion", f"Deleted students: {names} (IDs: {ids})", username)
        flash(f"Successfully deleted {len(ids)} students.", "danger")
        
    elif action == "status_update" and status_val:
        cur.execute("UPDATE students SET status = %s WHERE id = ANY(%s)", (status_val, ids))
        for sid in ids:
            cur.execute("INSERT INTO student_timeline (student_id, event_type, description) VALUES (%s, 'Status Update', %s)", (sid, f"Status updated via bulk action to '{status_val}'"))
        conn.commit()
        log_activity("Bulk Status Update", f"Updated status of student IDs {ids} to '{status_val}'.", username)
        flash(f"Successfully updated status of {len(ids)} students to {status_val}.", "success")

    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('home'))

# ==========================================
# DEPARTMENT MANAGEMENT (HOD OR ADMIN ONLY FOR MUTATION)
# ==========================================
@app.route("/departments")
@login_required
def departments_dashboard():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT d.id, d.name, d.code, d.description, 
               COUNT(s.id) AS student_count, 
               ROUND(COALESCE(AVG(s.gpa), 0), 2) AS avg_gpa
        FROM departments d
        LEFT JOIN students s ON s.department_id = d.id
        GROUP BY d.id, d.name, d.code, d.description
        ORDER BY d.name
    """)
    departments = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("departments.html", departments=departments)

@app.route("/departments/add", methods=["POST"])
@login_required
@hod_or_admin_required
def add_department():
    username = session['user']['username']
    name = request.form["name"].strip()
    code = request.form["code"].strip().upper()
    description = request.form.get("description", "").strip()
    
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO departments (name, code, description)
            VALUES (%s, %s, %s)
        """, (name, code, description))
        conn.commit()
        log_activity("Department Created", f"Created department '{name}' ({code})", username)
        flash(f"Department '{name}' created successfully!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to create department: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for('departments_dashboard'))

@app.route("/departments/edit/<int:id>", methods=["POST"])
@login_required
@hod_or_admin_required
def edit_department(id):
    username = session['user']['username']
    name = request.form["name"].strip()
    code = request.form["code"].strip().upper()
    description = request.form.get("description", "").strip()
    
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE departments
            SET name = %s, code = %s, description = %s
            WHERE id = %s
        """, (name, code, description, id))
        conn.commit()
        log_activity("Department Updated", f"Updated department ID: {id} to '{name}' ({code})", username)
        flash(f"Department '{name}' updated successfully!", "info")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to update department: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for('departments_dashboard'))

@app.route("/departments/delete/<int:id>")
@login_required
@hod_or_admin_required
def delete_department(id):
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT name FROM departments WHERE id = %s", (id,))
        dept_name = cur.fetchone()[0]
        cur.execute("DELETE FROM departments WHERE id = %s", (id,))
        conn.commit()
        log_activity("Department Deleted", f"Deleted department '{dept_name}' (ID: {id})", username)
        flash(f"Department '{dept_name}' deleted successfully!", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to delete department: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for('departments_dashboard'))

# ==========================================
# FACULTY ANNOUNCEMENTS HUB
# ==========================================
@app.route("/announcements")
@login_required
def announcements_dashboard():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.id, a.title, a.description, a.category, a.priority, a.created_at, a.expiry_date, u.username, a.scheduled_date
        FROM announcements a
        LEFT JOIN users u ON a.author_id = u.id
        ORDER BY a.created_at DESC
    """)
    announcements = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("announcements.html", announcements=announcements)

@app.route("/announcements/add", methods=["POST"])
@login_required
def add_announcement():
    username = session['user']['username']
    author_id = session['user']['id']
    title = request.form["title"].strip()
    description = request.form["description"].strip()
    category = request.form["category"].strip()
    priority = request.form["priority"].strip()
    scheduled_val = request.form.get("scheduled_date", "")
    expiry_val = request.form.get("expiry_date", "")

    scheduled_date = scheduled_val if scheduled_val else None
    expiry_date = expiry_val if expiry_val else None

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO announcements (title, description, category, priority, scheduled_date, expiry_date, author_id)
            VALUES (%s, %s, %s, %s, COALESCE(%s, CURRENT_TIMESTAMP), %s, %s)
            RETURNING id;
        """, (title, description, category, priority, scheduled_date, expiry_date, author_id))
        conn.commit()
        
        cur.execute("SELECT id FROM users")
        faculty_ids = [row[0] for row in cur.fetchall()]
        for fid in faculty_ids:
            create_notification(fid, 'Announcement', f"New notice: {title}")
            
        log_activity("Announcement Created", f"Posted announcement: {title}", username)
        flash("Announcement published successfully!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to post announcement: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for('announcements_dashboard'))

@app.route("/announcements/delete/<int:id>")
@login_required
def delete_announcement(id):
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM announcements WHERE id = %s", (id,))
        conn.commit()
        log_activity("Announcement Deleted", f"Deleted announcement ID: {id}", username)
        flash("Announcement removed successfully.", "danger")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to delete announcement: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for('announcements_dashboard'))

# ==========================================
# ANNOUNCEMENTS API ENDPOINTS
# ==========================================
@app.route("/api/announcements")
@login_required
def api_get_announcements():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT a.id, a.title, a.description, a.category, a.priority, a.created_at, a.expiry_date, u.username, a.scheduled_date
        FROM announcements a
        LEFT JOIN users u ON a.author_id = u.id
        ORDER BY a.created_at DESC
    """)
    announcements = cur.fetchall()
    cur.close()
    conn.close()
    
    result = []
    for row in announcements:
        result.append({
            'id': row[0],
            'title': row[1],
            'description': row[2],
            'category': row[3],
            'priority': row[4],
            'created_at': row[5].strftime('%Y-%m-%d %H:%M') if row[5] else None,
            'expiry_date': row[6].strftime('%Y-%m-%d %H:%M') if row[6] else None,
            'author': row[7] or 'System',
            'scheduled_date': row[8].strftime('%Y-%m-%d %H:%M') if row[8] else None,
            'is_author': (row[7] == session['user']['username'] or session['user']['role'] == 'Admin')
        })
    return jsonify({'status': 'success', 'announcements': result})

@app.route("/api/announcements/add", methods=["POST"])
@login_required
def api_add_announcement():
    username = session['user']['username']
    author_id = session['user']['id']
    
    if request.is_json:
        data = request.json
    else:
        data = request.form
        
    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    category = data.get("category", "General").strip()
    priority = data.get("priority", "Low").strip()
    scheduled_val = data.get("scheduled_date", "")
    expiry_val = data.get("expiry_date", "")

    if not title or not description:
        return jsonify({'status': 'error', 'message': 'Title and description are required.'}), 400

    scheduled_date = scheduled_val if scheduled_val else None
    expiry_date = expiry_val if expiry_val else None

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO announcements (title, description, category, priority, scheduled_date, expiry_date, author_id)
            VALUES (%s, %s, %s, %s, COALESCE(%s, CURRENT_TIMESTAMP), %s, %s)
            RETURNING id;
        """, (title, description, category, priority, scheduled_date, expiry_date, author_id))
        new_id = cur.fetchone()[0]
        conn.commit()
        
        cur.execute("SELECT id FROM users")
        faculty_ids = [row[0] for row in cur.fetchall()]
        for fid in faculty_ids:
            create_notification(fid, 'Announcement', f"New notice: {title}")
            
        log_activity("Announcement Created", f"Posted announcement: {title}", username)
        return jsonify({
            'status': 'success', 
            'message': 'Announcement published successfully!',
            'announcement': {
                'id': new_id,
                'title': title,
                'description': description,
                'category': category,
                'priority': priority,
                'created_at': 'Just now',
                'author': username
            }
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': f'Failed to post announcement: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/announcements/delete/<int:id>", methods=["POST"])
@login_required
def api_delete_announcement(id):
    username = session['user']['username']
    
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT author_id FROM announcements WHERE id = %s", (id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'status': 'error', 'message': 'Announcement not found.'}), 404
            
        author_id = row[0]
        if author_id != session['user']['id'] and session['user']['role'] != 'Admin':
            return jsonify({'status': 'error', 'message': 'You do not have permission to delete this announcement.'}), 403

        cur.execute("DELETE FROM announcements WHERE id = %s", (id,))
        conn.commit()
        log_activity("Announcement Deleted", f"Deleted announcement ID: {id}", username)
        return jsonify({'status': 'success', 'message': 'Announcement removed successfully.'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': f'Failed to delete announcement: {str(e)}'}), 500
    finally:
        cur.close()
        conn.close()


# ==========================================
# SYSTEM NOTIFICATIONS ENDPOINTS
# ==========================================
@app.route("/notifications/mark-read/<int:id>", methods=["POST"])
@login_required
def mark_notification_read(id):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE notifications SET is_read = TRUE WHERE id = %s AND user_id = %s", (id, session['user']['id']))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success'})

@app.route("/notifications/clear-all", methods=["POST"])
@login_required
def clear_all_notifications():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM notifications WHERE user_id = %s", (session['user']['id'],))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'status': 'success'})

# ==========================================
# PROFESSIONAL REPORTS HUB
# ==========================================
@app.route("/reports")
@login_required
def reports_hub():
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT COUNT(*) FROM students")
    total_students = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(*) FROM students WHERE status = 'Active'")
    active_students = cur.fetchone()[0]
    
    cur.execute("SELECT ROUND(AVG(gpa), 2) FROM students")
    avg_gpa_val = cur.fetchone()[0]
    avg_gpa = float(avg_gpa_val) if avg_gpa_val is not None else 0.00

    cur.execute("""
        SELECT d.name, d.code, COUNT(s.id) AS student_count, 
               ROUND(COALESCE(AVG(s.gpa), 0), 2) AS avg_gpa
        FROM departments d
        LEFT JOIN students s ON s.department_id = d.id
        GROUP BY d.id, d.name, d.code
        ORDER BY d.name
    """)
    departments = cur.fetchall()

    cur.execute("""
        SELECT s.id, s.name, s.email, d.name, s.gpa, s.status, s.age
        FROM students s
        LEFT JOIN departments d ON s.department_id = d.id
        ORDER BY d.name, s.name
    """)
    students = cur.fetchall()
    
    cur.close()
    conn.close()
    
    log_activity("Report Generated", "Compiled academic summary report.", username)
    return render_template("reports.html", 
                           total_students=total_students, 
                           active_students=active_students, 
                           avg_gpa=avg_gpa, 
                           departments=departments, 
                           students=students,
                           faculty_name=username,
                           current_date=time.strftime('%B %d, %Y'))

# ==========================================
# FACULTY MANAGEMENT HUB (ADMIN ONLY)
# ==========================================
@app.route("/faculty")
@login_required
@admin_required
def faculty_management():
    conn = get_connection()
    cur = conn.cursor()
    
    # Query all faculty requests and users
    cur.execute("""
        SELECT u.id, u.username, u.email, u.full_name, u.employee_id, d.name, u.designation, u.status, u.role, u.image_path, u.created_at, u.last_login
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        ORDER BY u.status DESC, u.created_at DESC
    """)
    faculty_list = cur.fetchall()
    
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cur.fetchall()
    
    cur.close()
    conn.close()
    return render_template("faculty_mgmt.html", faculty_list=faculty_list, departments=departments)

@app.route("/faculty/status/<int:fid>/<string:new_status>", methods=["POST"])
@login_required
@admin_required
def change_faculty_status(fid, new_status):
    admin_username = session['user']['username']
    if new_status not in ['Active', 'Rejected', 'Suspended']:
        flash("Invalid status request.", "danger")
        return redirect(url_for('faculty_management'))
        
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT username, email FROM users WHERE id = %s", (fid,))
    user_record = cur.fetchone()
    
    if not user_record:
        cur.close()
        conn.close()
        flash("Faculty record not found.", "warning")
        return redirect(url_for('faculty_management'))
        
    uname, uemail = user_record
    cur.execute("UPDATE users SET status = %s WHERE id = %s", (new_status, fid))
    conn.commit()
    cur.close()
    conn.close()

    # Log actions
    action_type = f"Faculty {new_status}"
    log_activity(action_type, f"Account '{uname}' status changed to {new_status} by Admin.", admin_username)
    
    # Send Notification to HOD/Faculty members
    create_notification(fid, 'Security', f"Your Faculty account status has been updated to {new_status}.")

    flash(f"Faculty account status updated to {new_status}!", "success")
    return redirect(url_for('faculty_management'))

@app.route("/faculty/role/<int:fid>", methods=["POST"])
@login_required
@admin_required
def change_faculty_role(fid):
    admin_username = session['user']['username']
    new_role = request.form["role"].strip()
    
    if new_role not in ['Admin', 'HOD', 'Faculty']:
        flash("Invalid role selection.", "danger")
        return redirect(url_for('faculty_management'))
        
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE users SET role = %s WHERE id = %s", (new_role, fid))
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Faculty Role Altered", f"Faculty ID {fid} role updated to {new_role}.", admin_username)
    create_notification(fid, 'System Alert', f"Your administrative role has been updated to {new_role}.")
    
    flash("Faculty role updated successfully.", "success")
    return redirect(url_for('faculty_management'))

@app.route("/faculty/reset-password/<int:fid>", methods=["POST"])
@login_required
@admin_required
def reset_faculty_password(fid):
    admin_username = session['user']['username']
    new_password = request.form["new_password"].strip()
    confirm_password = request.form["confirm_password"].strip()

    if new_password != confirm_password:
        flash("Passwords do not match.", "danger")
        return redirect(url_for('faculty_management'))

    is_valid_pwd, error_pwd = is_secure_password(new_password)
    if not is_valid_pwd:
        flash(error_pwd, "danger")
        return redirect(url_for('faculty_management'))

    hashed_pwd = generate_password_hash(new_password)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (hashed_pwd, fid))
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Faculty Password Reset", f"Administrative password override executed for user ID {fid}.", admin_username)
    create_notification(fid, 'Security', "Your password was updated by the system administrator.")
    
    flash("Faculty password updated successfully.", "success")
    return redirect(url_for('faculty_management'))

@app.route("/faculty/delete/<int:fid>", methods=["POST"])
@login_required
@admin_required
def delete_faculty(fid):
    admin_username = session['user']['username']
    
    # Protect deleting oneself
    if fid == session['user']['id']:
        flash("Cannot delete your own admin account.", "warning")
        return redirect(url_for('faculty_management'))

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT username FROM users WHERE id = %s", (fid,))
    res = cur.fetchone()
    uname = res[0] if res else f"ID: {fid}"
    
    cur.execute("DELETE FROM users WHERE id = %s", (fid,))
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Faculty Account Deleted", f"Deleted faculty account: {uname} (ID: {fid})", admin_username)
    flash(f"Faculty record '{uname}' has been deleted.", "danger")
    return redirect(url_for('faculty_management'))

# ==========================================
# FACULTY PROFILE & SETTINGS PAGE
# ==========================================
@app.route("/profile")
@login_required
def settings_page():
    username = session['user']['username']
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT u.username, u.email, u.full_name, u.employee_id, d.name, u.designation, u.image_path, u.role, u.status, u.created_at, u.last_login, u.department_id
        FROM users u
        LEFT JOIN departments d ON u.department_id = d.id
        WHERE u.username = %s
    """, (username,))
    user_info = cur.fetchone()
    
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cur.fetchall()
    
    cur.close()
    conn.close()
    return render_template("faculty_profile.html", user_info=user_info, departments=departments)

@app.route("/profile/update", methods=["POST"])
@login_required
def update_profile():
    current_username = session['user']['username']
    current_uid = session['user']['id']
    
    full_name = request.form["full_name"].strip()
    new_username = request.form["username"].strip()
    new_email = request.form["email"].strip().lower()
    designation = request.form["designation"].strip()
    department_id = int(request.form["department_id"])

    # Validate username
    is_valid, error_msg = is_secure_username(new_username)
    if not is_valid:
        flash(error_msg, "danger")
        return redirect(url_for('settings_page'))
        
    conn = get_connection()
    cur = conn.cursor()

    try:
        # Check duplicate username
        cur.execute("SELECT COUNT(*) FROM users WHERE username = %s AND id != %s", (new_username, current_uid))
        if cur.fetchone()[0] > 0:
            flash(f"Username '{new_username}' is already taken.", "danger")
            return redirect(url_for('settings_page'))

        # Check duplicate email
        cur.execute("SELECT COUNT(*) FROM users WHERE email = %s AND id != %s", (new_email, current_uid))
        if cur.fetchone()[0] > 0:
            flash(f"Email '{new_email}' is already taken.", "danger")
            return redirect(url_for('settings_page'))

        # Handle Profile photo upload
        image_path = None
        if 'profile_photo' in request.files:
            file = request.files['profile_photo']
            if file and file.filename != '' and allowed_file(file.filename):
                filename = f"avatar_{current_uid}_{int(time.time())}_{secure_filename(file.filename)}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = f"static/uploads/{filename}"

        if image_path:
            cur.execute("""
                UPDATE users 
                SET username = %s, email = %s, full_name = %s, designation = %s, department_id = %s, image_path = %s
                WHERE id = %s
            """, (new_username, new_email, full_name, designation, department_id, image_path, current_uid))
        else:
            cur.execute("""
                UPDATE users 
                SET username = %s, email = %s, full_name = %s, designation = %s, department_id = %s
                WHERE id = %s
            """, (new_username, new_email, full_name, designation, department_id, current_uid))
            
        conn.commit()

        # Update session
        session['user']['username'] = new_username
        session['user']['full_name'] = full_name

        log_activity("Faculty Profile Updated", f"User updated profile details.", new_username)
        create_notification(current_uid, 'Security', "Your profile details were updated.")
        flash("Profile updated successfully!", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Failed to update profile: {str(e)}", "danger")
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('settings_page'))

@app.route("/profile/change-password", methods=["POST"])
@login_required
def change_password():
    username = session['user']['username']
    uid = session['user']['id']
    current_password = request.form["current_password"].strip()
    new_password = request.form["new_password"].strip()
    confirm_password = request.form["confirm_password"].strip()

    if new_password != confirm_password:
        flash("New passwords do not match.", "danger")
        return redirect(url_for('settings_page'))

    is_valid_pwd, error_pwd = is_secure_password(new_password)
    if not is_valid_pwd:
        flash(error_pwd, "danger")
        return redirect(url_for('settings_page'))

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT password_hash FROM users WHERE id = %s", (uid,))
    record = cur.fetchone()
    
    if not record or not check_password_hash(record[0], current_password):
        cur.close()
        conn.close()
        flash("Incorrect current password.", "danger")
        return redirect(url_for('settings_page'))

    # Update
    new_password_hash = generate_password_hash(new_password)
    cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (new_password_hash, uid))
    conn.commit()
    cur.close()
    conn.close()

    log_activity("Password Changed", f"User changed password.", username)
    create_notification(uid, 'Security', "Your password was updated.")
    
    flash("Password updated successfully!", "success")
    return redirect(url_for('settings_page'))

# ==========================================
# ENTERPRISE ACADEMIC PLANNERS (CALENDAR)
# ==========================================
@app.route("/calendar")
@login_required
def calendar_dashboard():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM departments ORDER BY name")
    departments = cur.fetchall()
    
    cur.execute("""
        SELECT id, title, description, event_type, start_date, end_date, color_code
        FROM academic_events
        WHERE start_date >= CURRENT_DATE
        ORDER BY start_date ASC LIMIT 10
    """)
    upcoming_events = cur.fetchall()
    
    cur.close()
    conn.close()
    return render_template("calendar.html", departments=departments, upcoming_events=upcoming_events)

@app.route("/api/calendar/events")
@login_required
def api_get_calendar_events():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id, e.title, e.description, e.event_type, e.start_date, e.end_date, e.color_code, u.username
        FROM academic_events e
        LEFT JOIN users u ON e.faculty_id = u.id
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    events = []
    for r in rows:
        events.append({
            'id': r[0],
            'title': r[1],
            'description': r[2] or '',
            'type': r[3],
            'start': r[4].isoformat(),
            'end': r[5].isoformat(),
            'color': r[6],
            'author': r[7] or 'System'
        })
    return jsonify({'status': 'success', 'events': events})

@app.route("/api/calendar/events/add", methods=["POST"])
@login_required
def api_add_calendar_event():
    user_role = session['user']['role']
    faculty_id = session['user']['id']
    username = session['user']['username']
    
    data = request.json if request.is_json else request.form
    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    event_type = data.get("event_type", "Meeting")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    color_code = data.get("color_code", "#3b82f6")

    if user_role not in ['Admin', 'HOD'] and event_type in ['Exam', 'Holiday']:
        return jsonify({'status': 'error', 'message': 'Only HOD or Admin can create Exam/Holiday events.'}), 403

    if not title or not start_date or not end_date:
        return jsonify({'status': 'error', 'message': 'Title, start date, and end date are required.'}), 400

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO academic_events (title, description, event_type, start_date, end_date, color_code, faculty_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
        """, (title, description, event_type, start_date, end_date, color_code, faculty_id))
        new_id = cur.fetchone()[0]
        conn.commit()
        
        log_activity("Academic Event Created", f"Event: {title}", username)
        
        if event_type in ['Exam', 'Holiday']:
            cur.execute("SELECT id FROM users")
            all_uids = [row[0] for row in cur.fetchall()]
            for uid in all_uids:
                create_notification(uid, 'System Alert', f"New Academic Event: {title} scheduled.")
                
        return jsonify({'status': 'success', 'message': 'Event created successfully!', 'event_id': new_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/calendar/events/edit/<int:id>", methods=["POST"])
@login_required
def api_edit_calendar_event(id):
    user_role = session['user']['role']
    faculty_id = session['user']['id']
    username = session['user']['username']
    
    data = request.json if request.is_json else request.form
    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    event_type = data.get("event_type", "Meeting")
    start_date = data.get("start_date")
    end_date = data.get("end_date")
    color_code = data.get("color_code", "#3b82f6")

    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT faculty_id FROM academic_events WHERE id = %s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'Event not found.'}), 404
        
    creator_id = row[0]
    if creator_id != faculty_id and user_role != 'Admin':
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'You do not have permission to modify this event.'}), 403

    try:
        cur.execute("""
            UPDATE academic_events
            SET title = %s, description = %s, event_type = %s, start_date = %s, end_date = %s, color_code = %s
            WHERE id = %s
        """, (title, description, event_type, start_date, end_date, color_code, id))
        conn.commit()
        log_activity("Academic Event Modified", f"Event ID: {id}", username)
        return jsonify({'status': 'success', 'message': 'Event modified successfully!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/calendar/events/delete/<int:id>", methods=["POST"])
@login_required
def api_delete_calendar_event(id):
    user_role = session['user']['role']
    faculty_id = session['user']['id']
    username = session['user']['username']

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT faculty_id FROM academic_events WHERE id = %s", (id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'Event not found.'}), 404
        
    creator_id = row[0]
    if creator_id != faculty_id and user_role != 'Admin':
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'Permission denied.'}), 403

    try:
        cur.execute("DELETE FROM academic_events WHERE id = %s", (id,))
        conn.commit()
        log_activity("Academic Event Deleted", f"Event ID: {id}", username)
        return jsonify({'status': 'success', 'message': 'Event deleted successfully.'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==========================================
# BULK STUDENT IMPORT (EXCEL / CSV)
# ==========================================
@app.route("/api/students/import", methods=["POST"])
@login_required
def api_import_students():
    import csv
    import io
    import openpyxl
    
    user_role = session['user']['role']
    faculty_id = session['user']['id']
    username = session['user']['username']
    
    if user_role not in ['Admin', 'HOD']:
        return jsonify({'status': 'error', 'message': 'Access denied. Privileged role required.'}), 403
        
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file uploaded.'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'Empty file selected.'}), 400
        
    filename = secure_filename(file.filename)
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    if ext not in ['csv', 'xlsx']:
        return jsonify({'status': 'error', 'message': 'Unsupported file format. Please upload CSV or Excel (.xlsx).'}), 400
        
    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT email FROM students")
    existing_emails = {row[0].lower() for row in cur.fetchall() if row[0]}
    
    cur.execute("SELECT id, code FROM departments")
    dept_map = {row[1].upper(): row[0] for row in cur.fetchall()}
    
    imported_count = 0
    skipped_count = 0
    failed_count = 0
    errors_list = []
    
    try:
        cur.execute("""
            INSERT INTO student_imports (filename, faculty_id)
            VALUES (%s, %s) RETURNING id;
        """, (filename, faculty_id))
        import_id = cur.fetchone()[0]
    except Exception as e:
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': f'Failed to initialize import: {str(e)}'}), 500

    rows_to_process = []
    try:
        if ext == 'csv':
            stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
            csv_reader = csv.reader(stream)
            header = next(csv_reader, None)
            for idx, row in enumerate(csv_reader, 2):
                if len(row) >= 6:
                    rows_to_process.append((idx, row[0], row[1], row[2], row[3], row[4], row[5]))
        else:
            wb = openpyxl.load_workbook(file.stream)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if any(row):
                    row_data = [str(x).strip() if x is not None else '' for x in row]
                    if len(row_data) >= 6:
                        rows_to_process.append((idx, row_data[0], row_data[1], row_data[2], row_data[3], row_data[4], row_data[5]))
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': f'Failed to read import file structure: {str(e)}'}), 400

    seen_emails_in_batch = set()

    for row_num, name, age_str, dept_code, email, gpa_str, status in rows_to_process:
        name = name.strip()
        dept_code = dept_code.strip().upper()
        email = email.strip().lower()
        status = status.strip().title()
        
        if not name or not email:
            failed_count += 1
            reason = "Missing name or email"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))
            continue
            
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            failed_count += 1
            reason = "Invalid email format"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))
            continue

        try:
            age = int(age_str)
            if age <= 0:
                raise ValueError()
        except ValueError:
            failed_count += 1
            reason = "Invalid age (must be positive integer)"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))
            continue

        try:
            gpa = float(gpa_str)
            if gpa < 0.00 or gpa > 10.00:
                raise ValueError()
        except ValueError:
            failed_count += 1
            reason = "Invalid GPA (must be numeric 0.00 to 10.00)"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))
            continue

        if status not in ['Active', 'Inactive', 'Graduated']:
            failed_count += 1
            reason = "Invalid Status (must be Active, Inactive, or Graduated)"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))
            continue

        if dept_code not in dept_map:
            failed_count += 1
            reason = f"Department code '{dept_code}' not found"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))
            continue

        if email in existing_emails or email in seen_emails_in_batch:
            skipped_count += 1
            reason = "Duplicate record skipped (Email already exists)"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Skipped'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Skipped'))
            continue

        try:
            dept_id = dept_map[dept_code]
            cur.execute("""
                INSERT INTO students (name, age, department_id, email, gpa, status)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id;
            """, (name, age, dept_id, email, gpa, status))
            sid = cur.fetchone()[0]
            
            cur.execute("INSERT INTO student_marks (student_id, gpa) VALUES (%s, %s)", (sid, gpa))
            
            imported_count += 1
            seen_emails_in_batch.add(email)
        except Exception as insert_err:
            failed_count += 1
            reason = f"Database insertion error: {str(insert_err)}"
            errors_list.append({'row': row_num, 'name': name, 'email': email, 'reason': reason, 'status': 'Failed'})
            cur.execute("INSERT INTO student_import_errors (import_id, row_number, student_name, email, reason, status) VALUES (%s,%s,%s,%s,%s,%s)", (import_id, row_num, name, email, reason, 'Failed'))

    try:
        cur.execute("""
            UPDATE student_imports
            SET imported_count = %s, skipped_count = %s, failed_count = %s
            WHERE id = %s
        """, (imported_count, skipped_count, failed_count, import_id))
        conn.commit()
        log_activity("Bulk Import Executed", f"Import ID: {import_id}. Imported: {imported_count}, Skipped: {skipped_count}, Failed: {failed_count}", username)
    except Exception as summary_err:
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': f'Failed to complete import summary: {str(summary_err)}'}), 500

    cur.close()
    conn.close()
    
    return jsonify({
        'status': 'success',
        'imported': imported_count,
        'skipped': skipped_count,
        'failed': failed_count,
        'errors': errors_list
    })


# ==========================================
# ATTENDANCE MANAGEMENT MODULE
# ==========================================
@app.route("/attendance")
@login_required
def attendance_dashboard():
    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("SELECT id, name, code FROM departments ORDER BY name")
    departments = cur.fetchall()
    
    cur.execute("""
        SELECT s.id, s.name, d.name, s.email, d.id
        FROM students s
        LEFT JOIN departments d ON s.department_id = d.id
        WHERE s.status = 'Active'
        ORDER BY s.name
    """)
    students_list = cur.fetchall()
    
    cur.execute("""
        SELECT s.id, d.name, s.session_date, s.lecture_topic, u.username,
               COUNT(r.id) AS total_marked,
               COUNT(CASE WHEN r.status = 'Present' THEN 1 END) AS present_count
        FROM attendance_sessions s
        LEFT JOIN departments d ON s.department_id = d.id
        LEFT JOIN users u ON s.faculty_id = u.id
        LEFT JOIN attendance_records r ON r.session_id = s.id
        GROUP BY s.id, d.name, s.session_date, s.lecture_topic, u.username
        ORDER BY s.session_date DESC, s.created_at DESC LIMIT 20
    """)
    recent_sessions = cur.fetchall()
    
    cur.close()
    conn.close()
    return render_template("attendance.html", departments=departments, students_list=students_list, recent_sessions=recent_sessions)

@app.route("/api/attendance/session/create", methods=["POST"])
@login_required
def api_create_attendance_session():
    username = session['user']['username']
    faculty_id = session['user']['id']
    
    data = request.json if request.is_json else request.form
    dept_id = int(data.get("department_id"))
    session_date = data.get("session_date")
    lecture_topic = data.get("lecture_topic", "").strip()

    if not dept_id or not session_date or not lecture_topic:
        return jsonify({'status': 'error', 'message': 'Department, date, and lecture topic are required.'}), 400

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO attendance_sessions (department_id, session_date, lecture_topic, faculty_id)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (department_id, session_date, lecture_topic) 
            DO UPDATE SET lecture_topic = EXCLUDED.lecture_topic
            RETURNING id;
        """, (dept_id, session_date, lecture_topic, faculty_id))
        session_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'status': 'success', 'session_id': session_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/attendance/session/save", methods=["POST"])
@login_required
def api_save_attendance():
    username = session['user']['username']
    
    data = request.json
    session_id = int(data.get("session_id"))
    records = data.get("records", [])

    if not session_id or not records:
        return jsonify({'status': 'error', 'message': 'Invalid session parameters.'}), 400

    conn = get_connection()
    cur = conn.cursor()
    try:
        # Check if lecture_topic or session_date is passed in request, and update session details if so
        lecture_topic = data.get("lecture_topic")
        session_date = data.get("session_date")
        if lecture_topic or session_date:
            update_fields = []
            update_params = []
            if lecture_topic:
                update_fields.append("lecture_topic = %s")
                update_params.append(lecture_topic.strip())
            if session_date:
                update_fields.append("session_date = %s")
                update_params.append(session_date)
            update_params.append(session_id)
            cur.execute(f"""
                UPDATE attendance_sessions 
                SET {", ".join(update_fields)} 
                WHERE id = %s
            """, tuple(update_params))

        for r in records:
            sid = int(r.get("student_id"))
            stat = r.get("status")
            cur.execute("""
                INSERT INTO attendance_records (session_id, student_id, status)
                VALUES (%s, %s, %s)
                ON CONFLICT (session_id, student_id)
                DO UPDATE SET status = EXCLUDED.status;
            """, (session_id, sid, stat))
        conn.commit()
        log_activity("Attendance Marked", f"Session ID: {session_id}. Marked: {len(records)} records.", username)
        return jsonify({'status': 'success', 'message': 'Attendance saved successfully!'})
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/attendance/sessions")
@login_required
def api_get_attendance_sessions():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.id, d.name, s.session_date, s.lecture_topic, u.username,
               COUNT(r.id) AS total_marked,
               COUNT(CASE WHEN r.status = 'Present' THEN 1 END) AS present_count
        FROM attendance_sessions s
        LEFT JOIN departments d ON s.department_id = d.id
        LEFT JOIN users u ON s.faculty_id = u.id
        LEFT JOIN attendance_records r ON r.session_id = s.id
        GROUP BY s.id, d.name, s.session_date, s.lecture_topic, u.username
        ORDER BY s.session_date DESC, s.created_at DESC
    """)
    sessions = cur.fetchall()
    cur.close()
    conn.close()
    
    res = []
    for r in sessions:
        res.append({
            'id': r[0],
            'department_name': r[1],
            'session_date': r[2].isoformat() if r[2] else None,
            'lecture_topic': r[3],
            'faculty_username': r[4],
            'total_marked': r[5],
            'present_count': r[6]
        })
    return jsonify({'status': 'success', 'sessions': res})

@app.route("/api/attendance/session/<int:session_id>")
@login_required
def api_get_session_details(session_id):
    conn = get_connection()
    cur = conn.cursor()
    
    # Get session details
    cur.execute("""
        SELECT s.id, s.department_id, d.name, s.session_date, s.lecture_topic, u.username
        FROM attendance_sessions s
        LEFT JOIN departments d ON s.department_id = d.id
        LEFT JOIN users u ON s.faculty_id = u.id
        WHERE s.id = %s
    """, (session_id,))
    session_row = cur.fetchone()
    
    if not session_row:
        cur.close()
        conn.close()
        return jsonify({'status': 'error', 'message': 'Attendance session not found.'}), 404
        
    dept_id = session_row[1]
    
    # Get all active students of this department and their attendance status
    cur.execute("""
        SELECT s.id, s.name, s.email, COALESCE(r.status, 'Absent')
        FROM students s
        LEFT JOIN attendance_records r ON r.student_id = s.id AND r.session_id = %s
        WHERE s.department_id = %s AND s.status = 'Active'
        ORDER BY s.name
    """, (session_id, dept_id))
    students = cur.fetchall()
    cur.close()
    conn.close()
    
    student_list = []
    for row in students:
        student_list.append({
            'id': row[0],
            'name': row[1],
            'email': row[2],
            'status': row[3]
        })
        
    return jsonify({
        'status': 'success',
        'session': {
            'id': session_row[0],
            'department_id': session_row[1],
            'department_name': session_row[2],
            'session_date': session_row[3].isoformat() if session_row[3] else None,
            'lecture_topic': session_row[4],
            'faculty_username': session_row[5]
        },
        'students': student_list
    })


@app.route("/api/attendance/defaulters")
@login_required
def api_attendance_defaulters():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        WITH total_sessions AS (
            SELECT department_id, COUNT(id) AS total_count
            FROM attendance_sessions
            GROUP BY department_id
        ),
        student_presents AS (
            SELECT r.student_id, COUNT(r.id) AS present_count
            FROM attendance_records r
            WHERE r.status = 'Present'
            GROUP BY r.student_id
        )
        SELECT s.id, s.name, d.name AS dept_name, s.email, 
               COALESCE(ts.total_count, 0) AS total_classes, 
               COALESCE(sp.present_count, 0) AS attended_classes,
               ROUND(
                   CASE WHEN COALESCE(ts.total_count, 0) = 0 THEN 100.00
                   ELSE (COALESCE(sp.present_count, 0) * 100.00) / ts.total_count END,
               2) AS percentage
        FROM students s
        JOIN departments d ON s.department_id = d.id
        LEFT JOIN total_sessions ts ON ts.department_id = s.department_id
        LEFT JOIN student_presents sp ON sp.student_id = s.id
        WHERE s.status = 'Active'
        GROUP BY s.id, s.name, d.name, s.email, ts.total_count, sp.present_count
        HAVING (
            CASE WHEN COALESCE(ts.total_count, 0) = 0 THEN 100.00
            ELSE (COALESCE(sp.present_count, 0) * 100.00) / ts.total_count END
        ) < 75.00
        ORDER BY percentage ASC;
    """)
    defaulters = cur.fetchall()
    cur.close()
    conn.close()
    
    res = []
    for r in defaulters:
        res.append({
            'id': r[0],
            'name': r[1],
            'department': r[2],
            'email': r[3],
            'total': r[4],
            'present': r[5],
            'percentage': float(r[6])
        })
    return jsonify({'status': 'success', 'defaulters': res})

@app.route("/api/attendance/stats")
@login_required
def api_attendance_stats():
    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT COUNT(id), COUNT(CASE WHEN status='Present' THEN 1 END)
        FROM attendance_records
    """)
    row = cur.fetchone()
    total_recs = row[0] or 0
    present_recs = row[1] or 0
    overall_pct = round((present_recs * 100.00 / total_recs), 2) if total_recs > 0 else 100.00
    
    cur.execute("""
        SELECT d.name, COUNT(r.id), COUNT(CASE WHEN r.status='Present' THEN 1 END)
        FROM departments d
        JOIN students s ON s.department_id = d.id
        JOIN attendance_records r ON r.student_id = s.id
        GROUP BY d.id, d.name
    """)
    dept_stats = []
    for name, tot, pres in cur.fetchall():
        pct = round((pres * 100.00 / tot), 2) if tot > 0 else 100.00
        dept_stats.append({'department': name, 'percentage': float(pct)})
        
    cur.execute("""
        SELECT TO_CHAR(s.session_date, 'YYYY-MM') as month, COUNT(r.id), COUNT(CASE WHEN r.status='Present' THEN 1 END)
        FROM attendance_sessions s
        JOIN attendance_records r ON r.session_id = s.id
        GROUP BY month
        ORDER BY month
    """)
    months = []
    pcts = []
    for month, tot, pres in cur.fetchall():
        pct = round((pres * 100.00 / tot), 2) if tot > 0 else 100.00
        months.append(month)
        pcts.append(float(pct))
        
    cur.close()
    conn.close()
    return jsonify({
        'status': 'success',
        'overall': float(overall_pct),
        'departments': dept_stats,
        'trends': {
            'months': months,
            'percentages': pcts
        }
    })


# ==========================================
# MARKS & GPA MANAGEMENT MODULE
# ==========================================
@app.route("/marks")
@login_required
def marks_dashboard():
    conn = get_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT s.id, s.name, d.name, s.email,
               COALESCE(m.internal_marks, 0.00),
               COALESCE(m.lab_marks, 0.00),
               COALESCE(m.mid_marks, 0.00),
               COALESCE(m.final_marks, 0.00),
               COALESCE(m.total_marks, 0.00),
               COALESCE(m.percentage, 0.00),
               COALESCE(m.gpa, 0.00)
        FROM students s
        JOIN departments d ON s.department_id = d.id
        LEFT JOIN student_marks m ON m.student_id = s.id
        ORDER BY s.name
    """)
    marks_list = cur.fetchall()
    
    cur.execute("SELECT COUNT(*) FROM students")
    total_students = cur.fetchone()[0]
    
    cur.execute("""
        SELECT COUNT(*) FROM student_marks WHERE gpa >= 5.00
    """)
    passed_students = cur.fetchone()[0]
    
    pass_percentage = round((passed_students * 100.00 / total_students), 2) if total_students > 0 else 0.00
    
    cur.execute("""
        SELECT s.name, COALESCE(m.gpa, 0.00)
        FROM students s
        JOIN student_marks m ON m.student_id = s.id
        ORDER BY m.gpa DESC, s.name LIMIT 5
    """)
    top_performers = cur.fetchall()
    
    cur.close()
    conn.close()
    return render_template("marks.html", marks_list=marks_list, pass_percentage=pass_percentage, top_performers=top_performers, total_students=total_students)

@app.route("/api/marks/save", methods=["POST"])
@login_required
def api_save_marks():
    username = session['user']['username']
    
    data = request.json if request.is_json else request.form
    student_id = int(data.get("student_id"))
    internal = float(data.get("internal_marks", 0.00))
    lab = float(data.get("lab_marks", 0.00))
    mid = float(data.get("mid_marks", 0.00))
    final = float(data.get("final_marks", 0.00))

    if not (0.00 <= internal <= 20.00) or not (0.00 <= lab <= 20.00) or not (0.00 <= mid <= 30.00) or not (0.00 <= final <= 100.00):
        return jsonify({'status': 'error', 'message': 'Marks values exceed maximum allowed bounds.'}), 400

    total = internal + lab + mid + final
    percentage = round((total / 170.00) * 100.00, 2)
    gpa = round(percentage / 10.00, 2)

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO student_marks (student_id, internal_marks, lab_marks, mid_marks, final_marks, total_marks, percentage, gpa)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (student_id)
            DO UPDATE SET 
                internal_marks = EXCLUDED.internal_marks,
                lab_marks = EXCLUDED.lab_marks,
                mid_marks = EXCLUDED.mid_marks,
                final_marks = EXCLUDED.final_marks,
                total_marks = EXCLUDED.total_marks,
                percentage = EXCLUDED.percentage,
                gpa = EXCLUDED.gpa,
                updated_at = CURRENT_TIMESTAMP;
        """, (student_id, internal, lab, mid, final, total, percentage, gpa))
        
        cur.execute("""
            UPDATE students
            SET gpa = %s
            WHERE id = %s
        """, (gpa, student_id))
        
        conn.commit()
        log_activity("Marks Updated", f"Student ID: {student_id}. GPA: {gpa}", username)
        
        if gpa < 6.0:
            create_notification(session['user']['id'], 'Academic Alert', f"Student ID {student_id} is flagged at High Risk (GPA: {gpa})")
            
        return jsonify({
            'status': 'success',
            'message': 'Marks updated successfully!',
            'calculated': {
                'total': total,
                'percentage': percentage,
                'gpa': gpa
            }
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==========================================
# DEPARTMENT COMPARISON MODULE
# ==========================================
@app.route("/department-comparison")
@login_required
def department_comparison_dashboard():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, name, code FROM departments ORDER BY name")
    departments = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("department_comparison.html", departments=departments)

@app.route("/api/departments/compare")
@login_required
def api_compare_departments():
    conn = get_connection()
    cur = conn.cursor()
    
    ids_param = request.args.get("ids", "")
    if not ids_param:
        cur.execute("SELECT id FROM departments")
        dept_ids = [row[0] for row in cur.fetchall()]
    else:
        try:
            dept_ids = [int(x) for x in ids_param.split(",")]
        except ValueError:
            cur.close()
            conn.close()
            return jsonify({'status': 'error', 'message': 'Invalid department selections.'}), 400

    if not dept_ids:
        cur.close()
        conn.close()
        return jsonify({'status': 'success', 'comparison': []})

    placeholders = ",".join(["%s"] * len(dept_ids))
    query = f"""
        SELECT 
            d.id, 
            d.name, 
            d.code,
            COUNT(s.id) AS student_count,
            ROUND(COALESCE(AVG(s.gpa), 0.00), 2) AS avg_gpa,
            COUNT(CASE WHEN s.status = 'Active' THEN 1 END) AS active_count,
            COUNT(CASE WHEN s.status = 'Graduated' THEN 1 END) AS graduated_count,
            COUNT(CASE WHEN s.gpa >= 5.00 THEN 1 END) AS passed_count
        FROM departments d
        LEFT JOIN students s ON s.department_id = d.id
        WHERE d.id IN ({placeholders})
        GROUP BY d.id, d.name, d.code
        ORDER BY avg_gpa DESC
    """
    
    cur.execute(query, tuple(dept_ids))
    results = cur.fetchall()
    
    comparison = []
    for row in results:
        did = row[0]
        total_students = row[3]
        passed_students = row[7]
        pass_rate = round((passed_students * 100.00 / total_students), 2) if total_students > 0 else 0.00
        
        cur.execute("""
            SELECT name, gpa FROM students 
            WHERE department_id = %s AND status = 'Active'
            ORDER BY gpa DESC, name LIMIT 1
        """, (did,))
        top_row = cur.fetchone()
        top_perf = f"{top_row[0]} ({top_row[1]})" if top_row else "N/A"
        
        comparison.append({
            'id': row[0],
            'name': row[1],
            'code': row[2],
            'students': total_students,
            'avg_gpa': float(row[4]),
            'active': row[5],
            'graduated': row[6],
            'pass_rate': float(pass_rate),
            'top_performer': top_perf
        })
        
    cur.close()
    conn.close()
    return jsonify({'status': 'success', 'comparison': comparison})


# ==========================================
# SECURITY UTILITIES
# ==========================================
def is_secure_password(password):
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."
    if not re.search(r"[a-z]", password):
        return False, "Password must contain at least one lowercase letter."
    if not re.search(r"[A-Z]", password):
        return False, "Password must contain at least one uppercase letter."
    if not re.search(r"\d", password):
        return False, "Password must contain at least one digit."
    if not re.search(r"[@$!%*?&#]", password):
        return False, "Password must contain at least one special character (@$!%*?&#)."
    return True, ""

def is_secure_username(username):
    if len(username) < 3:
        return False, "Username must be at least 3 characters long."
    if not re.match(r"^[a-zA-Z0-9_\-]+$", username):
        return False, "Username can only contain alphanumeric characters, underscores, and hyphens."
    return True, ""

# ==========================================
# MAIN
# ==========================================
if __name__ == "__main__":
    app.run(debug=True)